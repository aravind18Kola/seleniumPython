import openpyxl

# for same data

# file = "C:\Sampless\Book2.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet1"]
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value = "Welcome"
# workbook.save(file)


# different data

file = "C:\Sampless\Book3.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "virat"

sheet.cell(2,1).value = 124
sheet.cell(2,2).value = "sachin"

sheet.cell(3,1).value = 125
sheet.cell(3,2).value = "viv"

workbook.save(file)
